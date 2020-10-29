from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Color, PatternFill
from meta import FILE_NAME_DISK
from CF_API import sort_data, data_all_contests, get_contest_name
import os
import openpyxl
import time

def clear_sheet():
    wb = load_workbook(f'data/{FILE_NAME_DISK}.xlsx')
    sheet = wb.active
    sheet.delete_cols(1, 100)
    sheet.delete_rows(1, 25)
    wb.save(f'data/{FILE_NAME_DISK}.xlsx')

def write_excel(data):
    if not os.path.exists(f'data/{FILE_NAME_DISK}.xlsx'):
        filepath = f"data/{FILE_NAME_DISK}.xlsx"
        wb = openpyxl.Workbook()
        wb.save(filepath)

    clear_sheet()

    wb = load_workbook(f'data/{FILE_NAME_DISK}.xlsx')
    sheet = wb.active
    free = 0
    Header = 3
    Side = 3

    #last_updated
    t = time.asctime(time.localtime(time.time()))
    sheet.cell(row=1, column=1).value = f'last_updated: {t}'

    #Side
    with open('data/nicknames.txt') as file:
        ind = 1
        for line in file:
            line = line.strip()
            sheet.cell(row=ind+Header, column=1).value = line
            ind += 1
    wb.save(f'data/{FILE_NAME_DISK}.xlsx')

    with open('data/contests_id.txt') as file:
        for id in file:
            id = int(id.strip())
            name_contest = get_contest_name(id)

            #Header
            sheet.cell(row=2, column=Side+free+1).value = name_contest
            for i in range(len(data[name_contest].columns)):
                sheet.cell(row=3, column=Side+free+1+i).value = data[name_contest].columns[i]


            for i in range(data[name_contest].shape[0]):
                for j in range(data[name_contest].shape[1]):
                    c = sheet.cell(row=Header+i + 1, column=free+Side+j+1)
                    Fill = PatternFill(patternType='lightTrellis',
                                          fgColor=Color('00AA00'))
                    if (data[name_contest].iat[i, j] == 1):
                        #c.fill = Fill
                        c.value = '+'

            free += data[name_contest].shape[1]
    wb.save(f'data/{FILE_NAME_DISK}.xlsx')






